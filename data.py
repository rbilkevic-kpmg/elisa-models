import openpyxl


class ElisaData:
    def __init__(self, elisa_data):
        self.elisa_data = elisa_data
        self._process()

    def _process(self):
        wb = openpyxl.load_workbook(filename=self.elisa_data, data_only=True)
        sheet = wb["End point_1"]

        tmp = [sheet.cell(row=row, column=col).value for col in range(2, 14) for row in range(31, 35)]

        self.STANDARDS_DATA = tmp[1:8]
        self.DATA = tmp[9:]

    def write_concentrations(self, conc_data):
        wb = openpyxl.load_workbook(filename=self.elisa_data)
        sheet = wb["End point_1"]

        conc_data = 9 * [None] + list(conc_data)
        for col in range(2, 14):
            for row in range(38, 42):
                sheet.cell(row=row, column=col).value = conc_data.pop(0)

        wb.save(self.elisa_data.replace('input\\', 'output\\').replace(".xlsx", "_result.xlsx"))
        print("Output saved in output folder")


class CRPData:
    def __init__(self, elisa_data):
        self.elisa_data = elisa_data
        self._process()

    def _process(self):
        wb = openpyxl.load_workbook(filename=self.elisa_data, data_only=True)
        sheet = wb["End point_1"]

        tmp = [sheet.cell(row=row, column=col).value for col in range(2, 14) for row in range(24, 28)]

        self.STANDARDS_DATA = [tmp[0]] + tmp[2:8]
        self.DATA = tmp[8:]

    def write_concentrations(self, conc_data):
        wb = openpyxl.load_workbook(filename=self.elisa_data)
        sheet = wb["End point_1"]

        conc_data = 8 * [None] + list(conc_data)
        for col in range(2, 14):
            for row in range(31, 35):
                sheet.cell(row=row, column=col).value = conc_data.pop(0)

        wb.save(self.elisa_data.replace('input\\', 'output\\').replace(".xlsx", "_result.xlsx"))
        print("Output saved in output folder")
